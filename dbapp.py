import os
import openpyxl
from openpyxl.styles import NamedStyle, Border, Side
import torch
from transformers import DistilBertTokenizerFast, DistilBertForTokenClassification
from transformers import DataCollatorForTokenClassification, Trainer, TrainingArguments
from datetime import datetime
from dateutil import parser
from dateutil.relativedelta import relativedelta


# Load DistilBERT tokenizer and model
model = DistilBertForTokenClassification.from_pretrained("./ner_model")
tokenizer = DistilBertTokenizerFast.from_pretrained("./ner_model")

# Get the label mappings from the model
id2label = {
    0: "O",
    1: "B-TASK",
    2: "I-TASK",
    3: "B-DEADLINE",
    4: "I-DEADLINE",
    5: "B-RECIPIENT",
    6: "I-RECIPIENT"
}
label2id = {v: k for k, v in id2label.items()}

def standardize_date(date_string):
    
    try:
        parsed_date = parser.parse(date_string, fuzzy=True)
        if parsed_date < datetime.now():
            parsed_date += relativedelta(years=1)
        return parsed_date.strftime("%d-%m-%Y")
    except ValueError as e:
        print(f"'{date_string}': {e}")
        return date_string
    
def check_and_create_file(file_path):
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(["Task", "Deadline", "Recipient", "Date Added"])
        workbook.save(file_path)
        print(f"Created new file: {file_path}")
    else:
        print(f"Writing to existing file: {file_path}")



def extract_tasks_and_deadlines(text, model, tokenizer):
    inputs = tokenizer(text, return_tensors="pt", padding=True, truncation=True)
    
    # Get model predictions
    with torch.no_grad():
        outputs = model(**inputs)
    
    # Get the predicted label ids
    predictions = torch.argmax(outputs.logits, dim=2)
    
    # Convert ids to labels
    predicted_labels = [id2label[p.item()] for p in predictions[0]]
    
    # Get the original tokens
    tokens = tokenizer.convert_ids_to_tokens(inputs['input_ids'][0])
    
    # Extract entities
    entities = []
    current_entity = None
    for token, label in zip(tokens, predicted_labels):
        if label != 'O':
            if current_entity is None:
                current_entity = {'type': label[2:], 'text': token}
            elif label[2:] == current_entity['type']:
                current_entity['text'] += ' ' + token
            else:
                entities.append(current_entity)
                current_entity = {'type': label[2:], 'text': token}
        else:
            if current_entity:
                entities.append(current_entity)
                current_entity = None
    
    if current_entity:
        entities.append(current_entity)
    
    # Clean up entity text
    for entity in entities:
        entity['text'] = entity['text'].replace(' ##', '').strip()
    
    return entities
def write_to_excel(file_path, tasks, deadlines, recipients):
   try:
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    
    date_style = NamedStyle(name='date_style')
    date_style.number_format = 'DD-MM-YYYY'
    date_style.border = Border(left=Side(style='thin'), 
                               right=Side(style='thin'), 
                               top=Side(style='thin'), 
                               bottom=Side(style='thin'))
    
    
    if date_style.name not in workbook.named_styles:
        workbook.add_named_style(date_style)
   
    next_row = worksheet.max_row + 1
    
    current_date = datetime.now().strftime("%d-%m-%Y")
    
    max_length = max(len(tasks), len(deadlines), len(recipients))
    
    for i in range(max_length):
        task = tasks[i] if i < len(tasks) else ""
        deadline = deadlines[i] if i < len(deadlines) else ""
        recipient = recipients[i] if i < len(recipients) else ""
       
        worksheet.cell(row=next_row, column=1).value = task
        deadline_cell = worksheet.cell(row=next_row, column=2)
        deadline_cell.value = standardize_date(deadline)
        deadline_cell.style = 'date_style' 
        worksheet.cell(row=next_row, column=3).value = recipient        
        date_cell = worksheet.cell(row=next_row, column=4)
        date_cell.value = current_date
        date_cell.style = 'date_style' 
        next_row += 1
       
        
    workbook.save(file_path)
   except Exception as e:
        print(f"An error occurred: {e}")
        
file_path = "tasks.xlsx"
check_and_create_file(file_path)
all_tasks = []
all_deadlines = []
all_recipients = []

while True:
    text_input = input("Enter a text (or type 'q' to finish): ")
    if text_input.lower() == 'q':
        break
    
    entities = extract_tasks_and_deadlines(text_input, model, tokenizer)
    tasks = [e['text'] for e in entities if e['type'] == 'TASK']
    deadlines = [e['text'] for e in entities if e['type'] == 'DEADLINE']
    recipients = [e['text'] for e in entities if e['type'] == 'RECIPIENT']
    concat_tasks = ", ".join(tasks)
    concat_deadlines = ", ".join(deadlines)
    concat_recipients = ", ".join(recipients)
    print(f"Extracted tasks: {tasks}")
    print(f"Extracted deadlines: {deadlines}")
    print(f"Extracted recipients: {recipients}")
    all_tasks.append(concat_tasks)
    all_deadlines.append(concat_deadlines)
    all_recipients.append(concat_recipients)


write_to_excel(file_path, all_tasks, all_deadlines, all_recipients)

print(f"Tasks and deadlines successfully extracted and saved to {file_path}")
